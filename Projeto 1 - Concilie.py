#importa as bibliotecas e funções que serão utilizadas.
from typing import Text
import PySimpleGUI as sg
from PySimpleGUI.PySimpleGUI import Print
from openpyxl import load_workbook, workbook
from funcoes import extrai_desc_aprop, extrai_desc_baixa, data
from datetime import datetime

#Gera a interface para visualização do usuário.
layout = [
            [sg.Text('Arquivo Razão:', size=(15,0)),sg.Input(size=(45,0),key='arquivo_razao')],
            [sg.Text('Arquivo Composição:', size=(15,0)),sg.Input(size=(45,0),key='arquivo_comp')],
            [sg.Text('Salvar em:', size=(15,0)),sg.Input(size=(45,0), key='salvar_em')],
            [sg.Text('Selecione abaixo o tipo de conciliação:')],            
            [sg.Radio('Fornecedor', 'tipo',key='tipo_fornecedor'), sg.Radio('Clientes', 'tipo', key='tipo_cliente'), sg.Radio('TicketLog','tipo',key='tipo_ticket')],
            [sg.Button('Integrar', size=(10,0)), sg.Button('Conciliar', size=(15,0)),sg.Button('Limpar', size=(10,0)), sg.Button('Cancelar', size=(10,0))],
            [sg.Output(size=(60,10))]
        ]

janela = sg.Window("Concilie!", layout)

while True:
    eventos, valores = janela.read()
    if eventos == 'Cancelar' or eventos == sg.WINDOW_CLOSED:
        break
    elif eventos == 'Limpar':
        janela['arquivo_razao']('')
        janela['arquivo_comp']('')
        janela['salvar_em']('')
    elif eventos == 'Integrar':
        arquivo_razao = valores['arquivo_razao']
        arquivo_comp = valores['arquivo_comp']
        salvar_em = valores['salvar_em']
        tipo_fornecedor = valores['tipo_fornecedor']
        tipo_cliente = valores['tipo_cliente']
        tipo_ticket = valores['tipo_ticket']

#verifica se é do tipo fornecedor
        if tipo_fornecedor == True:
# Atribui os arquivos e planilhas as variaveis.
            composicao = load_workbook(arquivo_comp)
            plan1_composicao = composicao.worksheets[0]
            plan2_baixas = composicao.worksheets[1]
            plan3_relatorio = composicao.worksheets[2]

            razao = load_workbook(arquivo_razao)
            plan1_razao = razao.worksheets[0]

            max_li_comp = plan1_composicao.max_row
            max_li_razao = plan1_razao.max_row
#inicio parte 1 - integrar dados
#Verifica se a descricao está de acordo com o padrão
            for i in range(2, max_li_razao + 1):
                texto = plan1_razao.cell(row=i, column=5).value
                provisao = 'PAGTO A EFETUAR DOC. NRO. '
                ver_prov = texto[0:26]
                if ver_prov != provisao:
                    baixa = 'PAGTO DO DOC. NRO. '
                    ver_bx = texto[0:19]
                    if ver_bx != baixa:
                        lista1 = []
                        dataf = data(plan1_razao.cell(row=i, column=2).value)
                        descricao = plan1_razao.cell(row=i, column=5).value
                        lcto = plan1_razao.cell(row=i, column=3).value
                        if plan1_razao.cell(row=i, column=7).value == 0:
                            valor = plan1_razao.cell(row=i, column=8).value
                        else:
                            valor = plan1_razao.cell(row=i, column=7).value
                            lista1.append(dataf)
                            lista1.append(lcto)
                            lista1.append(descricao)
                            lista1.append(valor)
                            lista1.append('Erro 01: A descrição está fora do padrão! Por favor, verifique.')
                            plan3_relatorio.append(lista1)

#faz a inclusão das linhas de provisão no arquivo de composiçção.
            i = 2
            while i < max_li_razao + 1:
                if plan1_razao.cell(row=i, column=7).value == 0:
                    desc = plan1_razao.cell(row=i, column=5).value #valor procurado
                    cod_lcto = plan1_razao.cell(row=i, column=3).value #codigo do valor procurado
                    soma = 0
                    cont = 0
                    lista2 = []
                    celula = 'E'
                    num_linha = 2
# verifica se a descrição se repete, percorrendo cada linha da planilha
                    while num_linha < max_li_razao + 1:
                        desc2 = plan1_razao[celula + str(num_linha)].value #valor que está sendo comparado
                        cod_lcto2 = plan1_razao['C' + str(num_linha)].value #codigo do valor que está sendo comparado
                        if desc == desc2:
                            if int(cod_lcto) == int(cod_lcto2) or int(cod_lcto2) == int(cod_lcto) + 1:
                                soma = soma + plan1_razao['H' + str(num_linha)].value
                                num_linha = num_linha + 1
                                cont = cont + 1
                            else:
                                lista3 = []
                                data_d = data(plan1_razao['B' + str(num_linha)].value)
                                lcto_d = plan1_razao['C' + str(num_linha)].value
                                descricao_d = plan1_razao['E' + str(num_linha)].value
                                valor_d = plan1_razao['H' + str(num_linha)].value
                                lista3.append(data_d)
                                lista3.append(lcto_d)
                                lista3.append(descricao_d)
                                lista3.append(valor_d)
                                lista3.append("Erro 02: Duplicidade! Por favor, verifique.")
                                plan3_relatorio.append(lista3)
                                num_linha = num_linha + 1
                                i = i + 1
                        else:
                            num_linha = num_linha + 1
                    if cont > 1:
                        data1 = data(plan1_razao.cell(row=i, column=2).value)
                        descricao = extrai_desc_aprop(plan1_razao.cell(row=i, column=5).value)
                        lista2.append(data1)
                        lista2.append(descricao)
                        lista2.append(soma)
                        plan1_composicao.append(lista2)
                        i = i + cont
                    else:
                        data1 = data(plan1_razao.cell(row=i, column=2).value)
                        descricao = extrai_desc_aprop(plan1_razao.cell(row=i, column=5).value)
                        valor = plan1_razao.cell(row=i, column=8).value
                        lista2.append(data1)
                        lista2.append(descricao)
                        lista2.append(valor)
                        plan1_composicao.append(lista2)
                        i = i + 1
                else:
                    i = i + 1

            composicao.save(salvar_em)
            print('Dados integrados com sucesso.')
        elif tipo_cliente == True:
            print('aqui deveria selecionar o arquivo de clientes')
        else:
            print('aqui deveria selecionar o arquivo tipo ticket')
#Fim da parte 1
#Concilia os dados verificando se a descrição que vem da baixa é igual a descrição que está 
#na planilha razão, caso positivo, verifica se o valor é igual, menor ou maior entre as planilhas
#e reage com a baixa, ou o levantamento dos erros.
    else:
        arquivo_razao = valores['arquivo_razao']
        arquivo_comp = valores['arquivo_comp']
        salvar_em = valores['salvar_em']
        tipo_fornecedor = valores['tipo_fornecedor']
        tipo_cliente = valores['tipo_cliente']
        tipo_ticket = valores['tipo_ticket']

        composicao = load_workbook(salvar_em)
        plan1_composicao = composicao.worksheets[0]
        plan2_baixas = composicao.worksheets[1]
        plan3_relatorio = composicao.worksheets[2]

        razao = load_workbook(arquivo_razao)
        plan1_razao = razao.worksheets[0]

        max_li_comp = plan1_composicao.max_row
        max_li_razao = plan1_razao.max_row

        if tipo_fornecedor == True:
            for i in range(2, max_li_razao + 1):
                if plan1_razao.cell(row=i, column=7).value != 0:
                    baixa = extrai_desc_baixa(plan1_razao.cell(row=i, column=5).value)
                    celula = 'B'
                    num_linha = 2
                    controle = 0
                    while num_linha < max_li_comp + 1:
                        provisao = plan1_composicao[celula + str(num_linha)].value
                        if baixa == provisao:
                            max_li_inconsistencia = plan3_relatorio.max_row
                            celula = 'C'
                            vlr_bx = plan1_razao.cell(row=i, column=7).value
                            vlr_provisao = plan1_composicao[celula + str(num_linha)].value
                            if vlr_bx == vlr_provisao:
# retira o registro completo da planilha composição e leva para a planilha baixa
# volta para a planilha composição e exclui a linha vazia
                                lista = []
                                dataf = data(plan1_composicao['A' + str(num_linha)].value)
                                descricao = plan1_composicao['B' + str(num_linha)].value
                                valor = plan1_composicao['C' + str(num_linha)].value
                                lista.append(dataf)
                                lista.append('vazio')
                                lista.append(descricao)
                                lista.append(valor)
                                plan2_baixas.append(lista)
                                plan1_composicao.delete_rows(num_linha)
                            elif vlr_bx > vlr_provisao:
# verifica se a baixa é maior que a provisao
# caso seja, copia o registro e gera uma linha no relatório de inconsistencias
# com o valor da subtração
                                lista = []
                                dataf = data(plan1_composicao['A' + str(num_linha)].value)
                                descricao = plan1_composicao['B' + str(num_linha)].value
                                valor = vlr_provisao - vlr_bx
                                lista.append(dataf)
                                lista.append('vazio')
                                lista.append(descricao)
                                lista.append(valor)
                                lista.append(
                                    'Erro 03: Valor da baixa maior que valor da apropriação, verifique multa/juros!')
                                plan3_relatorio.append(lista)
                            else:
# verifica se o valor da provisao é maior que a baixa, caso seja,
# faz uma copia do registro e lança na planilha de baixa com o valor da baixa
                                lista = []
                                dataf = data(plan1_composicao['A' + str(num_linha)].value)
                                descricao = plan1_composicao['B' + str(num_linha)].value
                                valor = vlr_bx
                                lista.append(dataf)
                                lista.append('vazio')
                                lista.append(descricao)
                                lista.append(valor)
                                plan2_baixas.append(lista)
# mantem uma copia na planilha de composicao com o valor composicao (-) a baixa
                                valor_sub = vlr_provisao - vlr_bx
                                plan1_composicao['C' + str(num_linha)].value = valor_sub
                            num_linha = 1
                            controle = controle + 1
                        else:
                            num_linha = num_linha + 1
                    if controle == 0:
                        lista_bx = []
                        data_bx = data(plan1_razao.cell(row=i, column=2).value)
                        lcto_bx = plan1_razao.cell(row=i, column=3).value
                        descricao_bx = extrai_desc_baixa(plan1_razao.cell(row=i, column=5).value)
                        valor_bx = plan1_razao.cell(row=i, column=7).value
                        lista_bx.append(data_bx)
                        lista_bx.append(lcto_bx)
                        lista_bx.append(descricao_bx)
                        lista_bx.append(valor_bx)
                        lista_bx.append("Erro 04: Baixa não encontrada. Por favor, verifique a provisão.")
                        plan3_relatorio.append(lista_bx)

            total_linhas_comp = plan1_composicao.max_row
            total_linhas_baixas = plan2_baixas.max_row
            total_linhas_rel_incons = plan3_relatorio.max_row

            composicao.save(salvar_em)

            if (total_linhas_rel_incons - 1) > 0:
                print(f'Atenção!'
                    f'\nForam encontradas {total_linhas_rel_incons-1} inconsistências.'
                    f'\nPor favor, verifique!')
                janela['arquivo_razao']('')
                janela['arquivo_comp']('')
                janela['salvar_em']('')        
            else:
                print(f'Conciliação concluída com sucesso!')
                janela['arquivo_razao']('')
                janela['arquivo_comp']('')
                janela['salvar_em']('')        

        elif tipo_cliente == True:
            print('Em desenvolvimento')
        else:
            print('Em desenvolvimento')